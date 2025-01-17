import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

if __name__ == '__main__':
	# 创建工作簿
	workbook = openpyxl.Workbook()

	# 创建 ABC 工作表，并设置相关格式
	item_sheet = workbook.active
	item_sheet.title = 'ABC'
	item_sheet.column_dimensions[get_column_letter(1)].width = 50
	blue_fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type='solid')
	center_align = Alignment(horizontal='center', vertical='center')
	white_font = Font(color='FFFFFFFF')
	blue_font = Font(color='FF000080', bold=True)
	item_sheet['A1'].fill = blue_fill
	item_sheet['A1'].font = white_font
	item_sheet['A1'].alignment = center_align
	item_sheet['A1'].value = '发布内容'
	item_sheet['B1'].fill = blue_fill
	item_sheet['B1'].value = '详情'
	item_sheet['C1'].value = '查看'
	item_sheet['C1'].font = blue_font
	item_sheet['C1'].hyperlink = "http://www.baidu.com"
	item_sheet.column_dimensions[get_column_letter(3)].width = 20

	# 创建 BCA 工作表，并设置相关格式
	bca_sheet = workbook.create_sheet('BCA')
	bca_sheet['A1'].value = '李四'

	# 保存工作簿
	workbook.save('./output.xlsx')

