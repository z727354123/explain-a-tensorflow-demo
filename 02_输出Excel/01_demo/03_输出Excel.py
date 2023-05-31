import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import json
import os


def createTable(workbook, item_sheet, title):
	if not workbook:
		workbook = openpyxl.Workbook()
		# 创建 ABC 工作表，并设置相关格式
		item_sheet = workbook.active
		item_sheet.title = title
	else:
		# 创建 BCA 工作表，并设置相关格式
		item_sheet = workbook.create_sheet(title)
	item_sheet.column_dimensions[get_column_letter(1)].width = 50
	blue_fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type='solid')
	center_align = Alignment(horizontal='center', vertical='center')
	white_font = Font(color='FFFFFFFF')
	blue_font = Font(color='FF000080', bold=True)
	item_sheet['A1'].fill = blue_fill
	item_sheet['A1'].font = white_font
	item_sheet['A1'].alignment = center_align
	item_sheet['A1'].value = '发布内容'
	item_sheet['B1'].fill = blue_fill
	item_sheet['B1'].font = white_font
	item_sheet['B1'].value = '详情'
	item_sheet.column_dimensions[get_column_letter(3)].width = 20
	return workbook, item_sheet


def setKeyVal(workbook, item_sheet, name, href, idx):
	blue_font = Font(color='FF000080', bold=True)
	AKey = f'A{idx}'
	BKey = f'B{idx}'
	item_sheet[AKey].value = name
	if href:
		item_sheet[BKey].value = '查看'
		item_sheet[BKey].font = blue_font
		item_sheet[BKey].hyperlink = href


def outFile(filename):
	json_file = f"./{filename}.json"
	workbook, item_sheet = None, None;
	if os.path.exists(json_file):  # json文件存在
		with open(json_file, encoding="utf-8") as f:
			json_data = json.load(f)  # JSON转字典
		for key, val in json_data.items():
			workbook, item_sheet = createTable(workbook, item_sheet, key)
			idx = 2
			for item in val:
				name = item.get("name")
				href = item.get("href")
				setKeyVal(workbook, item_sheet, name, href, idx)
				idx += 1;
	# 保存工作簿
	workbook.save(f'./{filename}.xlsx')


if __name__ == '__main__':
	# 创建工作簿
	# outFile('开发服务云')
	outFile('流程服务云')
	outFile('集成服务云')
	if True:
		exit(0)

